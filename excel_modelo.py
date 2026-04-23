import os
import time

import numpy as np
import urllib.request
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from recomendaciones import cfg_efectiva, recomendaciones_alerta, resumen_analisis

try:
    WINSOUND_OK = True
except ImportError:
    WINSOUND_OK = False

from tablas_calculos import MODEL_PATH, MODEL_URL, OBJECT_MODEL_PATH, OBJECT_MODEL_URL, NIVELES


def descargar_modelo(callback_status=None):
    try:
        if not os.path.exists(MODEL_PATH):
            if callback_status:
                callback_status("Descargando modelo de pose (~6 MB)...")
            urllib.request.urlretrieve(MODEL_URL, MODEL_PATH)
        if not os.path.exists(OBJECT_MODEL_PATH):
            if callback_status:
                callback_status("Descargando detector de objetos (~6 MB)...")
            urllib.request.urlretrieve(OBJECT_MODEL_URL, OBJECT_MODEL_PATH)
        return True
    except Exception as e:
        print(f"[modelo] Error al descargar: {e}")
        return False


HEADER_COLS = [
    ("Timestamp", 19),
    ("Trabajador", 22),
    ("Proyecto", 24),
    ("ROSA Final", 11),
    ("Nivel", 18),
    ("Total Silla", 12),
    ("Tabla B", 10),
    ("Tabla C", 10),
    ("Tabla D", 10),
    ("A1-Altura", 11),
    ("A2-Profund.", 12),
    ("A3-Reposabr.", 13),
    ("A4-Respaldo", 12),
    ("B1-Telefono", 12),
    ("B2-Pantalla", 12),
    ("C1-Mouse", 11),
    ("C2-Teclado", 12),
    ("F-Telefono", 11),
    ("F-Pantalla", 11),
    ("F-Mouse", 9),
    ("F-Teclado", 11),
    ("Ang.Tronco", 11),
    ("Ang.Rodilla", 11),
    ("Ang.Codo", 10),
    ("Desv.Cuello", 11),
    ("Ang.Muneca", 11),
]

COLOR_NIVEL = {1: "00D68F", 2: "00B87A", 3: "F6AD3B", 4: "E8823A", 5: "FF4757"}
UI_LABELS = {
    "pie_llega_suelo": "El pie llega al suelo",
    "altura_regulable": "Altura de la silla regulable",
    "espacio_insuficiente_piernas": "Espacio insuficiente para las piernas",
    "dist_rodilla_cm": "Distancia asiento-rodilla (cm)",
    "profundidad_regulable": "Profundidad del asiento regulable",
    "tiene_reposabrazos": "Tiene reposabrazos",
    "reposabrazos_ajustable": "Reposabrazos ajustable al codo",
    "reposabrazos_altos_bajos": "Reposabrazos altos o bajos",
    "bordes_afilados": "Reposabrazos con bordes duros",
    "brazos_anchos": "Reposabrazos demasiado anchos",
    "reposabrazos_no_regulables": "Reposabrazos no regulables",
    "usa_respaldo": "Usa el respaldo",
    "apoyo_lumbar_adecuado": "Apoyo lumbar adecuado",
    "hombros_encogidos_silla": "Hombros encogidos por mesa o silla",
    "respaldo_no_regulable": "Respaldo no regulable",
    "pantalla_dist_ok": "Pantalla entre 40 y 75 cm y a la altura de los ojos",
    "pantalla_baja": "Pantalla baja",
    "pantalla_elevada": "Pantalla elevada y obliga a extender el cuello",
    "dist_pantalla_mayor_75": "Pantalla a mas de 75 cm",
    "giro_otra_pantalla": "Gira el cuello para ver otra pantalla",
    "sin_portadocumentos": "No usa portadocumentos",
    "pantalla_reflejos": "Pantalla con reflejos",
    "raton_alineado": "Mouse alineado con el hombro y dentro del alcance",
    "agarre_pinza": "Agarre en pinza por mouse pequeno",
    "raton_teclado_dif_altura": "Mouse y teclado a distinta altura",
    "reposamanos_duro": "Reposamanos duro o con presion",
    "desviacion_escribir": "Desviacion al escribir",
    "alcance_sobre_cabeza": "Alcanza objetos por encima de la cabeza",
    "teclado_elevado_hombros": "Teclado elevado con hombros encogidos",
    "sin_soporte_teclado": "Teclado sin soporte ajustable",
    "telefono_alejado": "Telefono alejado del cuerpo (>30 cm) [base = 2]",
    "sujecion_hombro_cuello": "Sujecion del telefono con hombro/cuello [+2]",
    "sin_manos_libres": "No existe opcion de manos libres [+1]",
}


def _border():
    thin = Side(style="thin", color="C0C0C0")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _hdr_fill():
    return PatternFill("solid", fgColor="1F3864")


def _title_fill():
    return PatternFill("solid", fgColor="0A2342")


def nivel_texto(score):
    return NIVELES[min(max(score, 1), 5)][0]


def _meta_defaults(meta):
    meta = dict(meta or {})
    return {
        "nombre_proyecto": meta.get("nombre_proyecto", "Evaluacion ergonomica ROSA"),
        "evaluador": meta.get("evaluador", "Jaime Cesar Tarazona Tinoco"),
        "empresa": meta.get("empresa", "CHYC Ingenieros SAC"),
        "trabajador": meta.get("trabajador", "Trabajador"),
        "horas_diarias": meta.get("horas_diarias", ""),
        "camara": meta.get("camara", ""),
    }


def _cell(ws, row, col, value, *, bold=False, fill=None, color="000000", align="left"):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name="Arial", size=10, bold=bold, color=color)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.border = _border()
    if fill is not None:
        cell.fill = fill
    return cell


def _consideraciones_iniciales(cfg):
    filas = []
    for key, label in UI_LABELS.items():
        if key not in cfg:
            continue
        value = cfg[key]
        if isinstance(value, bool):
            estado = "Si" if value else "No"
        else:
            estado = value
        filas.append((label, estado))
    return filas


def _hoja_registros(wb):
    if "Registros ROSA" in wb.sheetnames:
        return wb["Registros ROSA"]
    ws = wb.active
    ws.title = "Registros ROSA"
    return ws


def _crear_hoja_resumen(wb):
    ws = wb.create_sheet("Resumen")
    ws.merge_cells("A1:D1")
    _cell(ws, 1, 1, "RESUMEN EVALUACION ROSA", bold=True, fill=_title_fill(), color="FFFFFF", align="center")
    for ci, titulo in enumerate(["Indicador", "Valor", "Descripcion", "Detalle"], 1):
        _cell(ws, 2, ci, titulo, bold=True, fill=_hdr_fill(), color="FFFFFF", align="center")
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 34
    ws.column_dimensions["D"].width = 48
    return ws


def _crear_hoja_consideraciones(wb, meta, cfg):
    ws = wb.create_sheet("Consideraciones UI")
    ws.merge_cells("A1:C1")
    _cell(ws, 1, 1, "CONSIDERACIONES INICIALES DE LA UI", bold=True, fill=_title_fill(), color="FFFFFF", align="center")
    meta_rows = [
        ("Proyecto", meta["nombre_proyecto"]),
        ("Evaluador", meta["evaluador"]),
        ("Empresa", meta["empresa"]),
        ("Trabajador", meta["trabajador"]),
        ("Horas diarias", meta["horas_diarias"]),
        ("Camara", meta["camara"]),
        ("Tabla F", ">4 h = +1 | 1-4 h = 0 | <1 h = -1"),
    ]
    row = 3
    for label, value in meta_rows:
        _cell(ws, row, 1, label, bold=True, fill=PatternFill("solid", fgColor="D9EAF7"))
        _cell(ws, row, 2, value)
        row += 1
    row += 1
    _cell(ws, row, 1, "Consideracion", bold=True, fill=_hdr_fill(), color="FFFFFF", align="center")
    _cell(ws, row, 2, "Valor inicial", bold=True, fill=_hdr_fill(), color="FFFFFF", align="center")
    row += 1
    for label, value in _consideraciones_iniciales(cfg):
        _cell(ws, row, 1, label)
        _cell(ws, row, 2, value, align="center")
        row += 1
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 18
    return ws


def _crear_hoja_informe(wb):
    ws = wb.create_sheet("Informe Proyecto")
    ws.merge_cells("A1:D1")
    _cell(ws, 1, 1, "INFORME DEL PROYECTO", bold=True, fill=_title_fill(), color="FFFFFF", align="center")
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 42
    return ws


def _actualizar_portada(ws, meta):
    total_cols = len(HEADER_COLS)
    fin = get_column_letter(total_cols)
    ws.merge_cells(f"A1:{fin}1")
    ws.merge_cells(f"A2:{fin}2")
    ws.merge_cells(f"A3:{fin}3")
    _cell(ws, 1, 1, "EVALUACION ERGONOMICA - METODO ROSA (NTP 1173 INSST 2022)", bold=True, fill=_title_fill(), color="FFFFFF", align="center")
    subtitulo = (
        f"Proyecto: {meta['nombre_proyecto']} | Empresa: {meta['empresa']} | "
        f"Evaluador: {meta['evaluador']}"
    )
    _cell(ws, 2, 1, subtitulo, fill=PatternFill("solid", fgColor="0D1B2A"), color="DDE7F5", align="center")
    detalle = (
        f"Trabajador: {meta['trabajador']} | Horas/dia: {meta['horas_diarias']} | "
        f"Camara: {meta['camara']} | Tabla F: >4 h = +1, 1-4 h = 0, <1 h = -1"
    )
    _cell(ws, 3, 1, detalle, fill=PatternFill("solid", fgColor="10253D"), color="DDE7F5", align="center")
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 20
    for ci, (titulo, ancho) in enumerate(HEADER_COLS, 1):
        cell = ws.cell(row=4, column=ci, value=titulo)
        cell.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        cell.fill = _hdr_fill()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _border()
        ws.column_dimensions[get_column_letter(ci)].width = ancho
    ws.row_dimensions[4].height = 34
    ws.freeze_panes = "A5"


def crear_excel_nuevo(ruta, metadata=None, cfg_inicial=None):
    meta = _meta_defaults(metadata)
    cfg = dict(cfg_inicial or {})
    wb = Workbook()
    ws = _hoja_registros(wb)
    _actualizar_portada(ws, meta)
    _crear_hoja_resumen(wb)
    _crear_hoja_consideraciones(wb, meta, cfg)
    _crear_hoja_informe(wb)
    wb.save(ruta)
    return wb


def _actualizar_resumen(wb, registro, num_fila):
    ws = wb["Registros ROSA"]
    ws2 = wb["Resumen"]
    fila_final = num_fila + 4
    rosas = []
    for r in range(5, fila_final + 1):
        value = ws.cell(row=r, column=4).value
        if isinstance(value, (int, float)):
            rosas.append(value)
    score = registro["rosa"]
    resumen_data = [
        ("Total registros", num_fila, "Evaluaciones realizadas", f"Ultima actualizacion: {registro.get('time', '')}"),
        ("Ultimo ROSA", score, "Resultado mas reciente", nivel_texto(score)),
        ("ROSA promedio", round(np.mean(rosas), 2) if rosas else "", "Media del periodo", ""),
        ("ROSA maximo", max(rosas) if rosas else "", "Puntuacion mas alta", ""),
        ("ROSA minimo", min(rosas) if rosas else "", "Puntuacion mas baja", ""),
        ("Alertas (>=5)", sum(1 for x in rosas if x >= 5), "Evaluaciones sobre nivel de accion", ""),
    ]
    for ri, (ind, val, desc, det) in enumerate(resumen_data, 3):
        _cell(ws2, ri, 1, ind, bold=True)
        _cell(ws2, ri, 2, val, align="center")
        _cell(ws2, ri, 3, desc)
        _cell(ws2, ri, 4, det)


def _actualizar_informe_proyecto(wb, registro, metadata, cfg_base=None):
    meta = _meta_defaults(metadata)
    ws = wb["Informe Proyecto"]
    for row in range(2, 80):
        for col in range(1, 5):
            ws.cell(row=row, column=col).value = None
            ws.cell(row=row, column=col).fill = PatternFill(fill_type=None)
            ws.cell(row=row, column=col).border = _border()
    filas_meta = [
        ("Nombre del proyecto", meta["nombre_proyecto"]),
        ("Evaluador del proyecto", meta["evaluador"]),
        ("Empresa", meta["empresa"]),
        ("Trabajador evaluado", meta["trabajador"]),
        ("Fecha de captura", registro.get("time", "")),
    ]
    row = 3
    for label, value in filas_meta:
        _cell(ws, row, 1, label, bold=True, fill=PatternFill("solid", fgColor="D9EAF7"))
        _cell(ws, row, 2, value)
        row += 1
    row += 1
    _cell(ws, row, 1, "Captura de tabla", bold=True, fill=_hdr_fill(), color="FFFFFF", align="center")
    _cell(ws, row, 2, "Valor", bold=True, fill=_hdr_fill(), color="FFFFFF", align="center")
    row += 1
    snapshot = [
        ("Tabla A", registro.get("tabla_A", "")),
        ("Tabla B", registro.get("tabla_B", "")),
        ("Tabla C", registro.get("tabla_C", "")),
        ("Tabla D", registro.get("tabla_D", "")),
        ("ROSA final", registro.get("rosa", "")),
    ]
    for label, value in snapshot:
        _cell(ws, row, 1, label)
        _cell(ws, row, 2, value, align="center")
        row += 1
    row += 1
    _cell(ws, row, 1, "Resultados analizados", bold=True, fill=_hdr_fill(), color="FFFFFF", align="center")
    _cell(ws, row + 1, 1, "Resumen")
    _cell(ws, row + 1, 2, resumen_analisis(registro, cfg_efectiva(cfg_base or {}, registro.get("flags_inferidos"))))
    row += 3
    cfg_eval = cfg_efectiva(cfg_base or {}, registro.get("flags_inferidos"))
    recomendaciones = recomendaciones_alerta(registro, cfg_eval)
    _cell(ws, row, 1, "Recomendaciones", bold=True, fill=_hdr_fill(), color="FFFFFF", align="center")
    _cell(ws, row, 2, "Acciones sugeridas", bold=True, fill=_hdr_fill(), color="FFFFFF", align="center")
    row += 1
    for item in recomendaciones:
        _cell(ws, row, 1, f"{item['categoria']} (puntaje {item['score']})", bold=True)
        _cell(ws, row, 2, "\n".join(f"- {accion}" for accion in item.get("acciones", [])))
        row += 1


def agregar_registro_excel(ruta, registro, num_fila, metadata=None, cfg_base=None):
    from openpyxl import load_workbook as _lw

    wb = _lw(ruta)
    meta = _meta_defaults(metadata)
    ws = wb["Registros ROSA"]
    _actualizar_portada(ws, meta)

    rosa = registro["rosa"]
    nivel = nivel_texto(rosa)
    color_nivel = COLOR_NIVEL.get(min(rosa, 5), "FF4757")
    fila = num_fila + 4
    valores = [
        registro.get("time", ""),
        meta["trabajador"],
        meta["nombre_proyecto"],
        rosa,
        nivel,
        registro.get("total_silla", ""),
        registro.get("tabla_B", ""),
        registro.get("tabla_C", ""),
        registro.get("tabla_D", ""),
        registro.get("A1", ""),
        registro.get("A2", ""),
        registro.get("A3", ""),
        registro.get("A4", ""),
        registro.get("B1", ""),
        registro.get("B2", ""),
        registro.get("C1", ""),
        registro.get("C2", ""),
        registro.get("factor_tiempo_telefono", ""),
        registro.get("factor_tiempo_pantalla", ""),
        registro.get("factor_tiempo_raton", ""),
        registro.get("factor_tiempo_teclado", ""),
        registro.get("ang_tronco", ""),
        registro.get("ang_rodilla", ""),
        registro.get("ang_codo", ""),
        registro.get("desv_cuello", ""),
        registro.get("ang_muneca", ""),
    ]
    fondo = "F0F6FF" if num_fila % 2 == 0 else "FFFFFF"
    for ci, val in enumerate(valores, 1):
        cell = ws.cell(row=fila, column=ci, value=val)
        cell.font = Font(name="Arial", size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _border()
        if ci in (4, 5):
            cell.fill = PatternFill("solid", fgColor=color_nivel)
            cell.font = Font(name="Arial", bold=True, size=10, color="FFFFFF" if rosa >= 3 else "000000")
        else:
            cell.fill = PatternFill("solid", fgColor=fondo)
    ws.row_dimensions[fila].height = 20

    _actualizar_resumen(wb, registro, num_fila)
    _actualizar_informe_proyecto(wb, registro, meta, cfg_base=cfg_base)
    wb.save(ruta)
